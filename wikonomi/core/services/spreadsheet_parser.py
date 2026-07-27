import csv
import io
import re
from decimal import Decimal, InvalidOperation
from pathlib import Path

from core.models import BulkImportRow
from core.services.bulk_config import MAX_ROWS


HEADER_ALIASES = {
    'sku': {'sku', 'product sku', 'item sku', 'item code', 'product code'},
    'barcode': {'barcode', 'bar code', 'ean', 'upc', 'gtin'},
    'product_name': {
        'product name',
        'product_name',
        'name',
        'item name',
        'description name',
    },
    'description': {'description', 'product description', 'details'},
    'brand': {'brand', 'manufacturer'},
    'category_name': {'category', 'category name', 'product category'},
    'unit': {'unit', 'uom', 'unit of measure', 'pack size'},
    'price': {'price', 'unit price', 'retail price', 'selling price'},
    'currency': {'currency', 'currency code'},
    'stock_quantity': {
        'stock quantity',
        'stock_quantity',
        'quantity',
        'qty',
        'stock',
        'on hand',
    },
    'notes': {'notes', 'note'},
    'tags': {'tags', 'tag'},
}

REQUIRED_FIELDS = {'product_name', 'price'}


def _clean_header(value):
    return re.sub(r'\s+', ' ', str(value or '').strip().lower().replace('_', ' '))


def canonical_header(value):
    cleaned = _clean_header(value)
    for canonical, aliases in HEADER_ALIASES.items():
        if cleaned in {_clean_header(alias) for alias in aliases}:
            return canonical
    return cleaned.replace(' ', '_')


def _clean_cell(value, limit=2000):
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).replace('\x00', '').strip()
    # Prevent spreadsheet formula execution in downloaded reports or later exports.
    if text.startswith(('=', '+', '-', '@')) and not re.fullmatch(r'-?\d+(\.\d+)?', text):
        text = "'" + text
    return text[:limit]


def _parse_decimal(value, *, positive=False):
    cleaned = re.sub(r'[^\d.\-]', '', _clean_cell(value, 80))
    if not cleaned:
        return None
    try:
        parsed = Decimal(cleaned)
    except (InvalidOperation, ValueError):
        return None
    if positive and parsed <= 0:
        return None
    return parsed


def _iter_csv(file_obj):
    raw = file_obj.read()
    for encoding in ('utf-8-sig', 'latin-1'):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError('File encoding is not supported. Save the CSV as UTF-8.')

    reader = csv.reader(io.StringIO(text))
    for row in reader:
        yield row


def _iter_xlsx(file_obj):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - deployment dependency guard
        raise ValueError('Excel support is not installed on the server.') from exc

    try:
        workbook = load_workbook(file_obj, read_only=True, data_only=True)
        sheet = workbook.active
        for row in sheet.iter_rows(values_only=True):
            yield list(row)
    except Exception as exc:
        raise ValueError('The Excel workbook is corrupt or could not be read.') from exc
    finally:
        try:
            workbook.close()
        except UnboundLocalError:
            pass


def _row_values(raw_row, headers):
    values = {}
    for index, header in enumerate(headers):
        values[header] = _clean_cell(raw_row[index] if index < len(raw_row) else '')
    return values


def parse_inventory_session(session):
    """Stream a CSV/XLSX into validated persistent rows."""
    extension = Path(session.inventory_filename).suffix.lower()
    iterator_factory = _iter_xlsx if extension == '.xlsx' else _iter_csv
    max_rows = MAX_ROWS

    session.rows.all().delete()
    created_rows = []
    missing_counts = {field: 0 for field in REQUIRED_FIELDS}
    seen_products = set()
    duplicate_products = 0

    with session.inventory_file.open('rb') as file_obj:
        iterator = iterator_factory(file_obj)
        try:
            raw_headers = next(iterator)
        except StopIteration:
            raise ValueError('The inventory spreadsheet is empty.')

        headers = [canonical_header(value) for value in raw_headers]
        missing_headers = REQUIRED_FIELDS - set(headers)
        if missing_headers:
            labels = ', '.join(sorted(field.replace('_', ' ') for field in missing_headers))
            raise ValueError(f'Missing required columns: {labels}.')

        for row_number, raw_row in enumerate(iterator, start=2):
            if row_number - 1 > max_rows:
                raise ValueError(f'This import exceeds the {max_rows:,}-row limit.')

            values = _row_values(raw_row, headers)
            if not any(values.values()):
                continue

            errors = []
            product_name = values.get('product_name', '')[:255]
            price = _parse_decimal(values.get('price'), positive=True)

            if not product_name:
                errors.append('Product name is required.')
                missing_counts['product_name'] += 1
            if len(values.get('product_name', '')) > 255:
                errors.append('Product name must be 255 characters or fewer.')
            if not values.get('price'):
                errors.append('Price is required.')
                missing_counts['price'] += 1
            elif price is None:
                errors.append('Price must be a positive number.')
            elif price > Decimal('9999999999.99'):
                errors.append('Price is above the supported maximum.')

            currency = values.get('currency', '').upper() or 'PGK'
            if not re.fullmatch(r'[A-Z]{3}', currency):
                errors.append('Currency must be a three-letter code such as PGK.')

            stock_quantity = None
            if values.get('stock_quantity'):
                stock_quantity = _parse_decimal(values.get('stock_quantity'))
                if stock_quantity is None or stock_quantity < 0:
                    errors.append('Stock quantity must be zero or a positive number.')

            duplicate_key = (
                values.get('sku', '').casefold()
                or values.get('barcode', '').casefold()
                or product_name.casefold()
            )
            if duplicate_key and duplicate_key in seen_products:
                duplicate_products += 1
                errors.append('Duplicate product row in this spreadsheet.')
            elif duplicate_key:
                seen_products.add(duplicate_key)

            created_rows.append(BulkImportRow(
                session=session,
                row_number=row_number,
                sku=values.get('sku', '')[:120],
                barcode=values.get('barcode', '')[:120],
                product_name=product_name,
                description=values.get('description', '')[:5000],
                brand=values.get('brand', '')[:255],
                category_name=values.get('category_name', '')[:255],
                unit=values.get('unit', '')[:80],
                price=price,
                currency=currency[:3],
                stock_quantity=stock_quantity,
                notes=values.get('notes', '')[:1000],
                tags=values.get('tags', '')[:1000],
                status=(
                    BulkImportRow.Status.INVALID
                    if errors
                    else BulkImportRow.Status.VALID
                ),
                validation_errors=errors,
            ))

            if len(created_rows) >= 1000:
                BulkImportRow.objects.bulk_create(created_rows, batch_size=1000)
                created_rows.clear()

    if created_rows:
        BulkImportRow.objects.bulk_create(created_rows, batch_size=1000)

    total_rows = session.rows.count()
    invalid_rows = session.rows.filter(status=BulkImportRow.Status.INVALID).count()
    valid_rows = total_rows - invalid_rows
    session.total_rows = total_rows
    session.valid_rows = valid_rows
    session.invalid_rows = invalid_rows
    session.duplicate_products = duplicate_products
    session.errors = [
        {
            'row': row.row_number,
            'errors': row.validation_errors,
        }
        for row in session.rows.filter(status=BulkImportRow.Status.INVALID)[:200]
    ]
    session.current_stage = 'Spreadsheet validation'
    session.progress_percent = 10 if invalid_rows == 0 else 0
    session.save(update_fields=[
        'total_rows',
        'valid_rows',
        'invalid_rows',
        'duplicate_products',
        'errors',
        'current_stage',
        'progress_percent',
        'updated_at',
    ])

    return {
        'total_rows': total_rows,
        'valid_rows': valid_rows,
        'invalid_rows': invalid_rows,
        'duplicate_products': duplicate_products,
        'missing_required_fields': missing_counts,
    }
