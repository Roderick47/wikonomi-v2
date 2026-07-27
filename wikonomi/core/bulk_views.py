import csv
from datetime import timedelta
from pathlib import Path

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.text import slugify
from django.views.decorators.http import require_POST

from core.models import (
    BulkImportImage,
    BulkImportRow,
    BulkImportSession,
    Business,
)
from core.services.image_matcher import ImageMatcher
from core.services.bulk_config import (
    MAX_IMAGE_SIZE,
    MAX_SPREADSHEET_SIZE,
    SESSION_DAYS,
)
from core.services.image_processor import (
    stage_uploaded_images,
    stage_zip_archive,
)
from core.services.inventory_importer import process_import_batch
from core.services.spreadsheet_parser import parse_inventory_session


INVENTORY_MIMES = {
    '.csv': {
        'text/csv',
        'application/csv',
        'application/vnd.ms-excel',
        'text/plain',
        'application/octet-stream',
    },
    '.xlsx': {
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'application/octet-stream',
        'application/zip',
    },
}


def _owned_session(request, session_id):
    return get_object_or_404(
        BulkImportSession.objects.select_related('business', 'user'),
        pk=session_id,
        user=request.user,
    )


def _unique_business_slug(name):
    base = slugify(name)[:220] or 'business'
    candidate = base
    counter = 2
    while Business.objects.filter(slug=candidate).exists():
        candidate = f'{base[:210]}-{counter}'
        counter += 1
    return candidate


def _resolve_business(request):
    business_id = request.POST.get('business', '').strip()
    if business_id.isdigit():
        return Business.objects.filter(pk=business_id).first()
    name = request.POST.get('business_name', '').strip()
    if not name:
        return None
    business = Business.objects.filter(name__iexact=name).first()
    if business:
        return business
    return Business.objects.create(name=name[:255], slug=_unique_business_slug(name))


def _inventory_file_error(uploaded):
    if uploaded is None:
        return 'Please select a CSV file or an Excel file to upload.'
    extension = Path(uploaded.name or '').suffix.lower()
    if extension not in INVENTORY_MIMES:
        return 'Inventory files must be CSV or XLSX.'
    max_size = MAX_SPREADSHEET_SIZE
    if uploaded.size > max_size:
        return f'Inventory file exceeds the {max_size // (1024 * 1024)} MB limit.'
    content_type = (getattr(uploaded, 'content_type', '') or '').lower()
    if content_type and content_type not in INVENTORY_MIMES[extension]:
        return 'The inventory file MIME type does not match its extension.'
    return None


@login_required
def inventory_upload(request):
    businesses = Business.objects.order_by('name')
    selected_business = None
    selected_id = request.GET.get('business', '')
    if selected_id.isdigit():
        selected_business = Business.objects.filter(pk=selected_id).first()

    active_session = (
        BulkImportSession.objects.filter(
            user=request.user,
            expires_at__gt=timezone.now(),
            status__in=[
                BulkImportSession.Status.PHOTOS,
                BulkImportSession.Status.REVIEW,
                BulkImportSession.Status.PROCESSING,
            ],
        )
        .select_related('business')
        .first()
    )
    active_session_url = ''
    if active_session:
        route_name = {
            BulkImportSession.Status.PHOTOS: 'bulk_upload_photos',
            BulkImportSession.Status.REVIEW: 'bulk_upload_review',
            BulkImportSession.Status.PROCESSING: 'bulk_upload_progress',
        }[active_session.status]
        active_session_url = reverse(
            route_name,
            kwargs={'session_id': active_session.pk},
        )

    context = {
        'wizard_step': 1,
        'businesses': businesses,
        'selected_business': selected_business,
        'active_session': active_session,
        'active_session_url': active_session_url,
    }
    if request.method == 'POST':
        uploaded = request.FILES.get('inventory_file')
        business = _resolve_business(request)
        error = _inventory_file_error(uploaded)
        if business is None:
            error = error or 'Choose or enter the business this inventory belongs to.'
        if error:
            messages.error(request, error)
            return render(request, 'bulk_inventory_wizard.html', context)

        import_session = BulkImportSession(
            user=request.user,
            business=business,
            inventory_filename=Path(uploaded.name).name[:255],
            expires_at=timezone.now() + timedelta(
                days=SESSION_DAYS
            ),
        )
        import_session.inventory_file = uploaded
        import_session.save()

        try:
            validation = parse_inventory_session(import_session)
        except ValueError as exc:
            import_session.errors = [{'row': 0, 'errors': [str(exc)]}]
            import_session.last_error = str(exc)
            import_session.save(update_fields=['errors', 'last_error', 'updated_at'])
            validation = {
                'total_rows': 0,
                'valid_rows': 0,
                'invalid_rows': 1,
                'missing_required_fields': {},
            }

        if validation['invalid_rows'] or not validation['valid_rows']:
            context.update({
                'import_session': import_session,
                'validation': validation,
                'validation_rows': import_session.rows.filter(
                    status=BulkImportRow.Status.INVALID
                )[:200],
                'selected_business': business,
            })
            messages.error(
                request,
                'Fix the spreadsheet errors before continuing to product photos.',
            )
            return render(request, 'bulk_inventory_wizard.html', context)

        import_session.status = BulkImportSession.Status.PHOTOS
        import_session.current_stage = 'Inventory validated'
        import_session.progress_percent = 15
        import_session.save(update_fields=[
            'status',
            'current_stage',
            'progress_percent',
            'updated_at',
        ])
        return redirect('bulk_upload_photos', session_id=import_session.pk)
    return render(request, 'bulk_inventory_wizard.html', context)


@login_required
def photo_upload(request, session_id):
    import_session = _owned_session(request, session_id)
    if import_session.invalid_rows:
        messages.error(request, 'The spreadsheet still contains validation errors.')
        return redirect('bulk_upload')
    if import_session.status == BulkImportSession.Status.COMPLETED:
        return redirect('bulk_upload_complete', session_id=import_session.pk)

    if request.method == 'POST':
        action = request.POST.get('action', 'upload')
        errors = []
        created_count = 0
        if action != 'skip':
            archive = request.FILES.get('photo_archive')
            photos = request.FILES.getlist('photos')
            if archive:
                if Path(archive.name or '').suffix.lower() != '.zip':
                    errors.append('Photo archive must be a ZIP file.')
                elif (getattr(archive, 'content_type', '') or '').lower() not in {
                    'application/zip',
                    'application/x-zip-compressed',
                    'application/octet-stream',
                }:
                    errors.append('The photo archive MIME type is not a ZIP.')
                else:
                    try:
                        created, archive_errors = stage_zip_archive(
                            import_session,
                            archive,
                        )
                        created_count += len(created)
                        errors.extend(archive_errors)
                    except ValueError as exc:
                        errors.append(str(exc))
            if photos:
                created, photo_errors = stage_uploaded_images(import_session, photos)
                created_count += len(created)
                errors.extend(photo_errors)
            if not archive and not photos:
                errors.append('Choose a ZIP, image files, or a folder before uploading.')

        import_session.warnings = list(import_session.warnings or []) + errors
        import_session.current_stage = 'Matching product photos'
        import_session.save(update_fields=['warnings', 'current_stage', 'updated_at'])
        ImageMatcher().match_session(import_session)
        review_url = reverse(
            'bulk_upload_review',
            kwargs={'session_id': import_session.pk},
        )
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({
                'created': created_count,
                'errors': errors,
                'review_url': review_url,
            }, status=200 if created_count or action == 'skip' else 400)
        for item in errors[:10]:
            messages.warning(request, item)
        return redirect(review_url)

    return render(request, 'bulk_inventory_photos.html', {
        'wizard_step': 2,
        'import_session': import_session,
        'existing_images': import_session.images.exclude(
            status=BulkImportImage.Status.REJECTED
        )[:100],
        'max_image_size_mb': MAX_IMAGE_SIZE // (1024 * 1024),
    })


def _manual_review_counts(import_session):
    matched_row_ids = set(
        import_session.images.filter(
            status__in=[
                BulkImportImage.Status.MATCHED,
                BulkImportImage.Status.IMPORTED,
            ],
            matched_row__isnull=False,
        ).values_list('matched_row_id', flat=True)
    )
    missing_rows = import_session.rows.filter(
        status__in=[
            BulkImportRow.Status.VALID,
            BulkImportRow.Status.IMPORTED,
        ],
    ).exclude(pk__in=matched_row_ids)
    return missing_rows


@login_required
def import_review(request, session_id):
    import_session = _owned_session(request, session_id)
    if import_session.status == BulkImportSession.Status.COMPLETED:
        return redirect('bulk_upload_complete', session_id=import_session.pk)

    if request.method == 'POST':
        for image in import_session.images.select_related('duplicate_of').all():
            match_value = request.POST.get(f'match_{image.pk}')
            if match_value:
                if match_value == 'skip':
                    image.status = BulkImportImage.Status.SKIPPED
                    image.matched_row = None
                elif match_value.isdigit():
                    row = import_session.rows.filter(pk=match_value).first()
                    if row:
                        image.matched_row = row
                        image.status = BulkImportImage.Status.MATCHED
                        image.match_method = 'manual'
                        image.confidence = 100
                image.save(update_fields=[
                    'matched_row',
                    'status',
                    'match_method',
                    'confidence',
                ])

            duplicate_action = request.POST.get(f'duplicate_{image.pk}')
            if duplicate_action == 'keep_both' and image.duplicate_of_id:
                image.matched_row = image.duplicate_of.matched_row
                image.duplicate_of = None
                image.status = (
                    BulkImportImage.Status.MATCHED
                    if image.matched_row_id
                    else BulkImportImage.Status.UPLOADED
                )
                image.save(update_fields=['matched_row', 'duplicate_of', 'status'])
            elif duplicate_action == 'keep_newest' and image.duplicate_of_id:
                older = image.duplicate_of
                older.status = BulkImportImage.Status.SKIPPED
                older.save(update_fields=['status'])
                image.matched_row = older.matched_row
                image.duplicate_of = None
                image.status = (
                    BulkImportImage.Status.MATCHED
                    if image.matched_row_id
                    else BulkImportImage.Status.UPLOADED
                )
                image.save(update_fields=['matched_row', 'duplicate_of', 'status'])
            elif duplicate_action == 'keep_oldest':
                image.status = BulkImportImage.Status.SKIPPED
                image.save(update_fields=['status'])

        for field_name, uploaded in request.FILES.items():
            if not field_name.startswith('missing_photo_'):
                continue
            row_id = field_name.removeprefix('missing_photo_')
            row = import_session.rows.filter(pk=row_id).first()
            if not row:
                continue
            created, errors = stage_uploaded_images(import_session, [uploaded])
            if created:
                image = created[0]
                image.matched_row = row
                image.status = BulkImportImage.Status.MATCHED
                image.match_method = 'manual_upload'
                image.confidence = 100
                image.save(update_fields=[
                    'matched_row',
                    'status',
                    'match_method',
                    'confidence',
                ])
            import_session.warnings = list(import_session.warnings or []) + errors

        matcher = ImageMatcher()
        matcher._assign_primary_images(import_session)
        matcher._update_session_counts(import_session)
        messages.success(request, 'Review choices saved.')
        return redirect('bulk_upload_review', session_id=import_session.pk)

    missing_rows = _manual_review_counts(import_session)
    return render(request, 'bulk_inventory_review.html', {
        'wizard_step': 4,
        'import_session': import_session,
        'unmatched_images': import_session.images.filter(
            status=BulkImportImage.Status.UNMATCHED
        ),
        'duplicate_images': import_session.images.filter(
            status=BulkImportImage.Status.DUPLICATE
        ).select_related('duplicate_of'),
        'missing_rows': missing_rows[:500],
        'missing_rows_count': missing_rows.count(),
        'warning_count': len(import_session.warnings or []),
        'rows': import_session.rows.filter(
            status=BulkImportRow.Status.VALID
        ).order_by('product_name'),
    })


@require_POST
@login_required
def staged_image_action(request, session_id, image_id):
    import_session = _owned_session(request, session_id)
    image = get_object_or_404(
        import_session.images,
        pk=image_id,
    )
    action = request.POST.get('action')
    if action == 'delete':
        image.file.delete(save=False)
        image.delete()
    elif action == 'skip':
        image.status = BulkImportImage.Status.SKIPPED
        image.matched_row = None
        image.save(update_fields=['status', 'matched_row'])
    ImageMatcher()._update_session_counts(import_session)
    return redirect('bulk_upload_review', session_id=import_session.pk)


@require_POST
@login_required
def start_import(request, session_id):
    import_session = _owned_session(request, session_id)
    if import_session.invalid_rows:
        messages.error(request, 'Spreadsheet validation must pass before import.')
        return redirect('bulk_upload')
    import_session.status = BulkImportSession.Status.PROCESSING
    import_session.current_stage = 'Preparing import'
    import_session.progress_percent = max(import_session.progress_percent, 40)
    import_session.started_at = import_session.started_at or timezone.now()
    import_session.last_error = ''
    import_session.save(update_fields=[
        'status',
        'current_stage',
        'progress_percent',
        'started_at',
        'last_error',
        'updated_at',
    ])
    return redirect('bulk_upload_progress', session_id=import_session.pk)


@login_required
def import_progress(request, session_id):
    import_session = _owned_session(request, session_id)
    if import_session.status == BulkImportSession.Status.COMPLETED:
        return redirect('bulk_upload_complete', session_id=import_session.pk)
    return render(request, 'bulk_inventory_progress.html', {
        'wizard_step': 5,
        'import_session': import_session,
    })


@require_POST
@login_required
def process_batch(request, session_id):
    try:
        with transaction.atomic():
            import_session = get_object_or_404(
                BulkImportSession.objects.select_for_update(),
                pk=session_id,
                user=request.user,
            )
            payload = process_import_batch(import_session)
    except Exception as exc:
        import_session = _owned_session(request, session_id)
        import_session.last_error = str(exc)
        import_session.current_stage = 'Paused — retrying is safe'
        import_session.save(update_fields=[
            'last_error',
            'current_stage',
            'updated_at',
        ])
        return JsonResponse({
            'status': import_session.status,
            'stage': import_session.current_stage,
            'progress_percent': import_session.progress_percent,
            'error': str(exc),
            'retryable': True,
        }, status=500)
    return JsonResponse(payload)


@login_required
def import_complete(request, session_id):
    import_session = _owned_session(request, session_id)
    if import_session.status != BulkImportSession.Status.COMPLETED:
        return redirect('bulk_upload_progress', session_id=import_session.pk)
    missing_count = _manual_review_counts(import_session).count()
    return render(request, 'bulk_inventory_complete.html', {
        'wizard_step': 6,
        'import_session': import_session,
        'missing_count': missing_count,
        'warning_count': len(import_session.warnings or []),
    })


def _safe_report_cell(value):
    text = str(value or '')
    return "'" + text if text.startswith(('=', '+', '-', '@')) else text


@login_required
def import_report(request, session_id):
    import_session = _owned_session(request, session_id)
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = (
        f'attachment; filename="wikonomi-import-{import_session.pk}.csv"'
    )
    writer = csv.writer(response)
    writer.writerow([
        'Filename',
        'Matched Product',
        'Match Method',
        'Confidence',
        'Status',
        'Warning',
        'Error',
    ])
    for image in import_session.images.select_related('matched_row').all():
        writer.writerow([
            _safe_report_cell(image.original_filename),
            _safe_report_cell(
                image.matched_row.product_name if image.matched_row else ''
            ),
            image.match_method,
            image.confidence,
            image.status,
            _safe_report_cell(image.warning),
            _safe_report_cell(image.error),
        ])
    for row in import_session.rows.filter(
        status__in=[BulkImportRow.Status.INVALID, BulkImportRow.Status.FAILED]
    ):
        writer.writerow([
            '',
            _safe_report_cell(row.product_name),
            '',
            '',
            row.status,
            '',
            _safe_report_cell('; '.join(row.validation_errors)),
        ])
    return response


@login_required
def download_inventory_template(request):
    response = HttpResponse(
        content_type=(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    )
    response['Content-Disposition'] = (
        'attachment; filename="wikonomi_inventory_template.xlsx"'
    )
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        messages.error(request, 'Excel template generation is unavailable.')
        return redirect('bulk_upload')

    workbook = Workbook(write_only=False)
    sheet = workbook.active
    sheet.title = 'Inventory'
    headers = [
        'SKU', 'Barcode', 'Product Name', 'Description', 'Brand', 'Category',
        'Unit', 'Price', 'Currency', 'Stock Quantity',
    ]
    sheet.append(headers)
    sheet.append([
        'RICE-10KG', '1234567890123', 'Rice 10kg', 'White rice bag',
        'Example Brand', 'Food', 'bag', 45.00, 'PGK', 25,
    ])
    for cell in sheet[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='4B2798')
    sheet.freeze_panes = 'A2'
    for column in sheet.columns:
        letter = column[0].column_letter
        sheet.column_dimensions[letter].width = min(
            max(len(str(cell.value or '')) for cell in column) + 2,
            35,
        )
    workbook.save(response)
    return response
